"""
[REP-04] CSV/XLSX reporter — genera export tabular de findings para tracking.
Produce dos archivos:
  - CSV: importable directo a Jira, ServiceNow, Excel, etc.
  - XLSX: versión formateada con auto-filtros, colores por severidad y hoja de resumen.
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reporter.base_reporter import BaseReporter
from schemas.controls_catalog import CONTROLS_BY_ID
from utils.logger import get_logger

logger = get_logger(__name__)

# ── Estilos XLSX ─────────────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SEV_FILLS = {
    "CRITICAL": PatternFill("solid", fgColor="FDE8E8"),
    "HIGH":     PatternFill("solid", fgColor="FEF0E5"),
    "MEDIUM":   PatternFill("solid", fgColor="FFF8E1"),
    "LOW":      PatternFill("solid", fgColor="E8F5E9"),
    "INFO":     PatternFill("solid", fgColor="E3F2FD"),
}

SEV_FONTS = {
    "CRITICAL": Font(name="Calibri", size=10, bold=True, color="7F0000"),
    "HIGH":     Font(name="Calibri", size=10, bold=True, color="C00000"),
    "MEDIUM":   Font(name="Calibri", size=10, bold=True, color="ED7D31"),
    "LOW":      Font(name="Calibri", size=10, bold=True, color="375623"),
    "INFO":     Font(name="Calibri", size=10, color="2E75B6"),
}

STATUS_FONTS = {
    "PASS": Font(name="Calibri", size=10, bold=True, color="375623"),
    "FAIL": Font(name="Calibri", size=10, bold=True, color="7F0000"),
    "SKIP": Font(name="Calibri", size=10, color="595959"),
}

PASS_FILL = PatternFill("solid", fgColor="E8F5E9")
FAIL_FILL = PatternFill("solid", fgColor="FDE8E8")
SKIP_FILL = PatternFill("solid", fgColor="F2F2F2")
STATUS_FILLS = {"PASS": PASS_FILL, "FAIL": FAIL_FILL, "SKIP": SKIP_FILL}

SUMMARY_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUMMARY_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

SEV_ORDER = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]

# ── Columnas de export ───────────────────────────────────────────────────────
CSV_HEADERS = [
    "control_id",
    "control_name",
    "service",
    "severity",
    "status",
    "message",
    "remediation",
    "framework",
]

XLSX_HEADERS = [
    "Control ID",
    "Control Name",
    "Service",
    "Severity",
    "Status",
    "Finding",
    "Remediation",
    "Framework",
]

XLSX_COL_WIDTHS = [18, 45, 14, 12, 10, 55, 55, 10]


class CsvReporter(BaseReporter):

    def generate(self) -> str:
        logger.info("Generating CSV/XLSX export...")

        csv_path  = self._output_path("csv")
        xlsx_path = self._output_path("xlsx")
        csv_path.parent.mkdir(parents=True, exist_ok=True)

        rows = self._build_rows()

        self._write_csv(csv_path, rows)
        self._write_xlsx(xlsx_path, rows)

        logger.info(f"CSV  export saved: {csv_path}")
        logger.info(f"XLSX export saved: {xlsx_path}")
        return str(xlsx_path)

    # ── Data preparation ─────────────────────────────────────────────────────

    def _build_rows(self) -> list[dict]:
        """Construye lista de dicts con todos los findings ordenados."""
        rows = []
        for f in sorted(self.report.findings, key=lambda x: (
            SEV_ORDER.index(x.severity.value) if x.severity.value in SEV_ORDER else 99,
            0 if x.status.value == "FAIL" else (1 if x.status.value == "SKIP" else 2),
            x.service,
        )):
            catalog = CONTROLS_BY_ID.get(f.control_id, {})
            framework = catalog.get("framework", "CIS" if f.control_id.startswith("CIS") else "WAF")
            rows.append({
                "control_id":   f.control_id,
                "control_name": f.control_name,
                "service":      f.service.upper(),
                "severity":     f.severity.value,
                "status":       f.status.value,
                "message":      f.message,
                "remediation":  f.remediation or "",
                "framework":    framework,
            })
        return rows

    # ── CSV ───────────────────────────────────────────────────────────────────

    def _write_csv(self, path: Path, rows: list[dict]):
        with open(path, "w", newline="", encoding="utf-8-sig") as fp:
            writer = csv.DictWriter(fp, fieldnames=CSV_HEADERS)
            writer.writeheader()
            writer.writerows(rows)

    # ── XLSX ──────────────────────────────────────────────────────────────────

    def _write_xlsx(self, path: Path, rows: list[dict]):
        wb = Workbook()

        # Sheet 1: Findings
        self._write_findings_sheet(wb, rows)

        # Sheet 2: Summary
        self._write_summary_sheet(wb)

        wb.save(str(path))

    def _write_findings_sheet(self, wb: Workbook, rows: list[dict]):
        ws = wb.active
        ws.title = "Findings"
        ws.sheet_properties.tabColor = "1F3864"

        # Freeze header row
        ws.freeze_panes = "A2"

        # Headers
        for col_idx, (header, width) in enumerate(zip(XLSX_HEADERS, XLSX_COL_WIDTHS), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        data_font = Font(name="Calibri", size=10)
        wrap_align = Alignment(vertical="top", wrap_text=True)

        for row_idx, row_data in enumerate(rows, 2):
            values = [row_data[k] for k in CSV_HEADERS]
            alt_fill = PatternFill("solid", fgColor="F8F9FA") if row_idx % 2 == 0 else None

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = wrap_align
                cell.border = THIN_BORDER

                # Background alterno
                if alt_fill:
                    cell.fill = alt_fill

            # Colorear columna Severity (col 4)
            sev_cell = ws.cell(row=row_idx, column=4)
            sev = row_data["severity"]
            if sev in SEV_FONTS:
                sev_cell.font = SEV_FONTS[sev]
            if sev in SEV_FILLS:
                sev_cell.fill = SEV_FILLS[sev]

            # Colorear columna Status (col 5)
            status_cell = ws.cell(row=row_idx, column=5)
            status = row_data["status"]
            if status in STATUS_FONTS:
                status_cell.font = STATUS_FONTS[status]
            if status in STATUS_FILLS:
                status_cell.fill = STATUS_FILLS[status]

        # Auto-filtro
        last_row = len(rows) + 1
        last_col = get_column_letter(len(XLSX_HEADERS))
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        # Row height
        ws.row_dimensions[1].height = 30

    def _write_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Summary")
        ws.sheet_properties.tabColor = "2E75B6"

        r = self.report
        current_row = 1

        # ── Title ────────────────────────────────────────────────────────────
        ws.merge_cells("A1:D1")
        title_cell = ws.cell(row=1, column=1, value="AWS Security Audit — Findings Summary")
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1F3864")
        current_row = 3

        # ── Metadata ─────────────────────────────────────────────────────────
        meta = [
            ("Account ID",   r.account_id),
            ("Region",       r.region),
            ("Generated",    r.generated_at.strftime('%Y-%m-%d %H:%M UTC')),
            ("Auditor",      r.auditor or "N/A"),
            ("Global Score", f"{r.global_score}%"),
        ]
        for label, value in meta:
            ws.cell(row=current_row, column=1, value=label).font = Font(name="Calibri", size=10, bold=True, color="1F3864")
            ws.cell(row=current_row, column=2, value=value).font = Font(name="Calibri", size=10)
            current_row += 1

        current_row += 1

        # ── Overall counts ───────────────────────────────────────────────────
        headers = ["Metric", "Count"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.fill = SUMMARY_HEADER_FILL
            cell.font = SUMMARY_HEADER_FONT
            cell.border = THIN_BORDER
        current_row += 1

        counts = [
            ("Total Controls", r.total_controls),
            ("Passed",         r.passed_controls),
            ("Failed",         r.failed_controls),
            ("Skipped",        r.skipped_controls),
        ]
        for label, val in counts:
            ws.cell(row=current_row, column=1, value=label).font = Font(name="Calibri", size=10)
            ws.cell(row=current_row, column=1).border = THIN_BORDER
            val_cell = ws.cell(row=current_row, column=2, value=val)
            val_cell.font = Font(name="Calibri", size=10, bold=True)
            val_cell.border = THIN_BORDER
            val_cell.alignment = Alignment(horizontal="center")
            current_row += 1

        current_row += 1

        # ── By severity ──────────────────────────────────────────────────────
        headers = ["Severity", "Count"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.fill = SUMMARY_HEADER_FILL
            cell.font = SUMMARY_HEADER_FONT
            cell.border = THIN_BORDER
        current_row += 1

        for sev in SEV_ORDER:
            count = r.by_severity.get(sev, 0)
            sev_cell = ws.cell(row=current_row, column=1, value=sev)
            sev_cell.border = THIN_BORDER
            if sev in SEV_FONTS:
                sev_cell.font = SEV_FONTS[sev]
            if sev in SEV_FILLS:
                sev_cell.fill = SEV_FILLS[sev]
            cnt_cell = ws.cell(row=current_row, column=2, value=count)
            cnt_cell.font = Font(name="Calibri", size=10, bold=True)
            cnt_cell.border = THIN_BORDER
            cnt_cell.alignment = Alignment(horizontal="center")
            current_row += 1

        current_row += 1

        # ── By service ───────────────────────────────────────────────────────
        svc_headers = ["Service", "Total", "Passed", "Failed", "Skipped", "Score"]
        for col_idx, h in enumerate(svc_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.fill = SUMMARY_HEADER_FILL
            cell.font = SUMMARY_HEADER_FONT
            cell.border = THIN_BORDER
        current_row += 1

        for s in r.by_service:
            vals = [s.service.upper(), s.total, s.passed, s.failed, s.skipped, f"{s.score}%"]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.border = THIN_BORDER
                if col_idx == 1:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="1F3864")
                elif col_idx == 4 and s.failed > 0:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="7F0000")
                elif col_idx == 6:
                    cell.alignment = Alignment(horizontal="center")
            current_row += 1

        # Column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12